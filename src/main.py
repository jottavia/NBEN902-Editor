import tkinter as tk
from tkinter import filedialog, messagebox
from editable_table import EditableTable
from file_operations import parse_fixed_width_file, save_data
from member_selector import MemberSelector
import os
import subprocess
import pandas as pd
from openpyxl import load_workbook
from sftp_transmitter import SFTPTransmitter, save_before_transmit

def main():
    root = tk.Tk()
    root.title("Fixed-Width File Editor")
    root.geometry("1000x600")

    file_path = None
    default_filename = "paysrp.nben902.sccea.input"
    columns = [
        ("Agency Code", 10),
        ("Name", 50),
        ("Employee ID", 10),
        ("Deduction Code", 6),
        ("Effective Date", 10),
        ("Deduction End Date", 10),
        ("Deduction Amount", 8)
    ]

    def open_file():
        nonlocal file_path
        new_file_path = filedialog.askopenfilename(initialfile=default_filename, filetypes=[("Input files", "*.input"), ("All files", "*.*")])
        if not new_file_path:
            return
        new_df = parse_fixed_width_file(new_file_path)
        current_data = editable_table.get_data()
        if current_data:
            current_df = pd.DataFrame(current_data, columns=[col[0] for col in columns])
            combined_df = pd.concat([current_df, new_df], ignore_index=True)
        else:
            combined_df = new_df
        editable_table.load_data(combined_df)
        file_path = new_file_path

    def save_file(save_file_path=None):
        nonlocal file_path
        if not save_file_path:
            save_file_path = filedialog.asksaveasfilename(initialfile=os.path.basename(file_path) if file_path else default_filename, defaultextension=".input", filetypes=[("Input files", "*.input"), ("All files", "*.*")])
            if not save_file_path:
                return
        data = editable_table.get_data()
        if data is None:
            return
        df = pd.DataFrame(data, columns=[col[0] for col in columns])
        save_data(df, save_file_path)
        file_path = save_file_path

    def add_new_row():
        editable_table.add_row()

    def modify_member():
        member_file_path = filedialog.askopenfilename(title="Select Member Data File", filetypes=[("Excel files", "*.xlsx")])
        if not member_file_path:
            return
        
        try:
            # Read the member data file
            wb = load_workbook(member_file_path)
            sheet = wb.active
            member_data = []
            for row in sheet.iter_rows(values_only=True):
                member_data.append(row)
            
            # Extract relevant data from the member_data list
            if member_data:
                headers = member_data[0]
                data = member_data[1:]
                df_member = pd.DataFrame(data, columns=headers)
                
                if df_member.empty:
                    messagebox.showerror("Error", "The member data file is empty.")
                    return
                
                # Ask the user to select members
                selector = MemberSelector(root, df_member)
                root.wait_window(selector)
                selected_members = selector.get_selected_members()
                
                for new_member_info in selected_members:
                    new_row = {}
                    for col in columns:
                        column_name = col[0]
                        if column_name in new_member_info:
                            new_row[column_name] = new_member_info[column_name]
                        elif column_name == "Deduction Code":
                            new_row[column_name] = "407"
                        elif column_name == "Deduction Amount":
                            new_row[column_name] = "100"  # Set deduction amount to 100
                        elif column_name == "Employee ID":
                            new_row[column_name] = new_member_info[headers[5]]  # Use the 6th column for Employee ID
                        else:
                            new_row[column_name] = ""
                    
                    editable_table.add_row(pd.Series(new_row))
        except Exception as e:
            messagebox.showerror("Error", f"Failed to read the member data file: {e}")

    def transmit_902():
        nonlocal file_path
        file_path = save_before_transmit(file_path, save_file)
        if file_path:
            SFTPTransmitter(root, file_path, save_file)
            
    def close_app():
        data = editable_table.get_data()
        if data is None:
            return
        new_df = pd.DataFrame(data, columns=[col[0] for col in columns])

        if file_path and not new_df.equals(parse_fixed_width_file(file_path)):
            response = messagebox.askyesnocancel("Save Changes", "Do you want to save changes before closing?")
            if response is None:
                return
            elif response:
                save_file()

        root.destroy()

    # Button Frame
    button_frame = tk.Frame(root)
    button_frame.grid(row=0, column=0, sticky="nw")

    open_button = tk.Button(button_frame, text="Open", command=open_file)
    open_button.pack(side="left", padx=(10, 5), pady=5)

    save_button = tk.Button(button_frame, text="Save", command=save_file)
    save_button.pack(side="left", padx=(0, 5), pady=5)

    add_row_button = tk.Button(button_frame, text="Add Row", command=add_new_row)
    add_row_button.pack(side="left", padx=(0, 5), pady=5)

    modify_member_button = tk.Button(button_frame, text="Modify Member", command=modify_member)
    modify_member_button.pack(side="left", padx=(0, 5), pady=5)

    transmit_902_button = tk.Button(button_frame, text="Transmit 902", command=transmit_902)
    transmit_902_button.pack(side="left", padx=(0, 5), pady=5)

    exit_button = tk.Button(button_frame, text="Exit", command=root.quit)
    exit_button.pack(side="left", padx=(0, 10), pady=5)

    # Table Frame
    table_frame = tk.Frame(root)
    table_frame.grid(row=1, column=0, sticky="nsew")

    root.grid_rowconfigure(1, weight=1)
    root.grid_columnconfigure(0, weight=1)

    canvas = tk.Canvas(table_frame)
    canvas.grid(row=0, column=0, sticky="nsew")

    table_frame.grid_rowconfigure(0, weight=1)
    table_frame.grid_columnconfigure(0, weight=1)

    editable_table = EditableTable(canvas, columns=columns)
    canvas.create_window((0, 0), window=editable_table, anchor="nw")

    def update_scrollregion(event):
        canvas.configure(scrollregion=canvas.bbox("all"))

    editable_table.bind("<Configure>", update_scrollregion)

    def on_mouse_wheel(event):
        canvas.yview_scroll(int(-1*(event.delta/120)), "units")

    canvas.bind_all("<MouseWheel>", on_mouse_wheel)

    scrollbar_y = tk.Scrollbar(table_frame, orient="vertical", command=canvas.yview)
    scrollbar_y.grid(row=0, column=1, sticky="ns")
    canvas.configure(yscrollcommand=scrollbar_y.set)

    scrollbar_x = tk.Scrollbar(root, orient="horizontal", command=canvas.xview)
    scrollbar_x.grid(row=2, column=0, sticky="ew")
    canvas.configure(xscrollcommand=scrollbar_x.set)

    root.protocol("WM_DELETE_WINDOW", close_app)
    root.mainloop()

if __name__ == "__main__":
    main()
